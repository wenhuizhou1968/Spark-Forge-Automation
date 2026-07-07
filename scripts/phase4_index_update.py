#!/usr/bin/env python3
"""
Phase 4: Index Update & Report Generation
Updates Living Index and generates biweekly report
"""

import os
import sys
import json
import logging
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/phase4_index_update.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

load_dotenv()
SPARK_FORGE_PATH = os.getenv('SPARK_FORGE_PATH', r'Z:\R&D\R&D Meeting\AI-Spark-Forge')

def load_all_entries():
    """Load entries from all phases"""
    logger.info("Loading all entries...")

    all_entries = {
        'spark_entries': [],
        'forge_entries': []
    }

    try:
        if Path('phase1_extracted_entries.json').exists():
            with open('phase1_extracted_entries.json', 'r') as f:
                phase1 = json.load(f)
                all_entries['spark_entries'].extend(phase1.get('spark_entries', []))
                all_entries['forge_entries'].extend(phase1.get('forge_entries', []))

        if Path('phase2_external_entries.json').exists():
            with open('phase2_external_entries.json', 'r') as f:
                phase2 = json.load(f)
                all_entries['spark_entries'].extend(phase2.get('spark_entries', []))
                all_entries['forge_entries'].extend(phase2.get('forge_entries', []))

        logger.info(f"Loaded {len(all_entries['spark_entries'])} Spark + {len(all_entries['forge_entries'])} Forge entries")

    except Exception as e:
        logger.error(f"Error loading entries: {e}")

    return all_entries

def load_analysis():
    """Load trend analysis from Phase 3"""
    logger.info("Loading trend analysis...")

    analysis = {
        'red_flags': [],
        'yellow_signals': [],
        'convergence_signals': {}
    }

    try:
        if Path('phase3_analysis.json').exists():
            with open('phase3_analysis.json', 'r') as f:
                analysis = json.load(f)

    except Exception as e:
        logger.error(f"Error loading analysis: {e}")

    return analysis

def update_living_index(spark_entries, forge_entries):
    """Update Excel Living Index with new entries"""
    logger.info("Updating Living Index...")

    index_path = Path(SPARK_FORGE_PATH) / 'Spark_Forge_Index.xlsx'

    try:
        # Create or load workbook
        if index_path.exists():
            wb = load_workbook(index_path)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Living Index'

            # Create headers
            headers = ['Entry #', 'Date', 'Type', 'Title', 'Presenter', 'Category', 'Status', 'Tags', 'Strategic Link', 'Key Insight', 'Link', 'Comment']
            ws.append(headers)
            start_row = 2

        # Add Spark entries
        for entry in spark_entries:
            ws.append([
                entry.get('entry_id'),
                entry.get('date'),
                'Spark',
                entry.get('title'),
                entry.get('presenter', entry.get('source', '')),
                entry.get('category'),
                'Active',
                ', '.join(entry.get('tags', [])),
                entry.get('strategic_link'),
                '',
                '',
                entry.get('source', '')
            ])

        # Add Forge entries
        for entry in forge_entries:
            ws.append([
                entry.get('entry_id'),
                entry.get('date'),
                'Forge',
                entry.get('title'),
                entry.get('presenter', entry.get('vendor', '')),
                entry.get('category'),
                'Active',
                ', '.join(entry.get('tags', [])),
                entry.get('strategic_link'),
                '',
                '',
                entry.get('source', '')
            ])

        # Save workbook
        Path(SPARK_FORGE_PATH).mkdir(parents=True, exist_ok=True)
        wb.save(index_path)
        logger.info(f"Updated Living Index: {index_path}")

    except Exception as e:
        logger.error(f"Error updating Living Index: {e}")

def generate_biweekly_report(spark_entries, forge_entries, analysis):
    """Generate biweekly trend report"""
    logger.info("Generating biweekly report...")

    try:
        # Calculate date range
        today = datetime.now()
        start_date = (today - datetime.timedelta(days=14)).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

        # Create report content
        report = f"""# Biweekly Spark & Forge Trend Report
**Report Date:** {start_date} – {end_date}
**Prepared by:** Spark_Forge Automation
**Distribution:** Wenhui Zhou + Group Leaders

---

## 📌 Executive Snapshot

| Metric | Count |
|--------|-------|
| New Spark entries | {len(spark_entries)} |
| New Forge entries | {len(forge_entries)} |
| RED flags detected | {len(analysis.get('red_flags', []))} |
| YELLOW signals | {len(analysis.get('yellow_signals', []))} |
| GREEN signals | {len(analysis.get('green_signals', []))} |

---

## 🔴 RED FLAGS (Immediate Action Required)

"""

        # Add RED flags
        red_flags = analysis.get('red_flags', [])
        if red_flags:
            for flag in red_flags[:5]:  # Top 5
                signal = analysis.get('convergence_signals', {}).get(flag, {})
                report += f"**{flag}** (Score: {signal.get('score')}, Sources: {', '.join(signal.get('sources', []))})\n\n"
        else:
            report += "*None detected this cycle.*\n\n"

        report += """---

## 🟡 YELLOW SIGNALS (Emerging Themes)

"""

        # Add YELLOW signals
        yellow_signals = analysis.get('yellow_signals', [])
        if yellow_signals:
            for signal in yellow_signals[:5]:
                report += f"- {signal}\n"
        else:
            report += "*None detected this cycle.*\n\n"

        report += f"""---

## 📊 New Entries This Cycle

### Spark Entries ({len(spark_entries)})

"""

        # Add Spark entries
        for entry in spark_entries[:10]:  # Show first 10
            report += f"**{entry.get('entry_id')}:** {entry.get('title')}\n"
            report += f"- Category: {entry.get('category')}\n"
            report += f"- Tags: {', '.join(entry.get('tags', []))}\n\n"

        report += f"""### Forge Entries ({len(forge_entries)})

"""

        # Add Forge entries
        for entry in forge_entries[:10]:  # Show first 10
            report += f"**{entry.get('entry_id')}:** {entry.get('title')}\n"
            report += f"- Category: {entry.get('category')}\n"
            report += f"- Tags: {', '.join(entry.get('tags', []))}\n\n"

        report += f"""---

## 📈 Analysis & Insights

- Total entries this cycle: {len(spark_entries) + len(forge_entries)}
- Convergence signals detected: {len(analysis.get('convergence_signals', {}))}
- Highest priority themes: {', '.join(red_flags[:3]) if red_flags else 'None'}

---

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""

        # Save report
        report_path = Path(SPARK_FORGE_PATH) / 'Biweekly_Reports' / f"{today.strftime('%Y-%m-%d')}_Trend_Report.md"
        report_path.parent.mkdir(parents=True, exist_ok=True)

        with open(report_path, 'w') as f:
            f.write(report)

        logger.info(f"Generated biweekly report: {report_path}")
        return str(report_path)

    except Exception as e:
        logger.error(f"Error generating report: {e}")
        return None

def main():
    """Main Phase 4 execution"""
    try:
        logger.info("=" * 60)
        logger.info("PHASE 4: Index Update & Report Generation")
        logger.info("=" * 60)

        # Load all entries
        all_entries = load_all_entries()
        spark_entries = all_entries['spark_entries']
        forge_entries = all_entries['forge_entries']

        # Load analysis
        analysis = load_analysis()

        # Update Living Index
        update_living_index(spark_entries, forge_entries)

        # Generate biweekly report
        report_path = generate_biweekly_report(spark_entries, forge_entries, analysis)

        logger.info("Phase 4 Complete ✓")
        logger.info(f"Files saved to: {SPARK_FORGE_PATH}")

    except Exception as e:
        logger.error(f"Phase 4 Failed: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
